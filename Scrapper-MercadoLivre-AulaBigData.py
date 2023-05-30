import requests
from bs4 import BeautifulSoup
import openpyxl
from getpass import getpass
from email.message import EmailMessage
import re
import yagmail  # Importando a biblioteca yagmail

class Scrappy: #Classe nomeada como Scrappy, ela sera responsavel por entrar em um website no caso o mercado livre, puxar dados de precos e nomes, salvar um arquivo Excel e enviar por email esse arquivo!

    def __init__(self): #Esse método é o construtor da classe Scrappy. Ele é chamado quando um novo objeto Scrappy é criado. Ele inicializa o email do usuário como None e cria duas listas vazias para armazenar os nomes e os preços dos produtos
        self.email = None
        self.product_name_list = []
        self.product_price_list = []

    def start(self):
        #Inicia o Script (Este método chama os outros métodos na ordem em que devem ser executados. Primeiro, obtém o email do usuário, então busca os dados do produto, cria o arquivo Excel e, finalmente, envia o email. Se algum erro ocorrer durante esse processo, ele será capturado e impresso.)
        try:
            self.get_user_email()
            self.scrape_data()
            self.create_excel_file()
            self.send_email()
        except Exception as e:
            print(f"An error occurred: {e}")

    def get_user_email(self):
        #Faz o prompt da primeira pergunta para o usuario no terminal, pedindo o email do usuario (Este método solicita ao usuário que digite um endereço de e-mail e verifica se o endereço de e-mail é válido usando uma expressão regular. Se o endereço de e-mail não for válido, ele solicitará que o usuário digite um novo endereço de e-mail.)
        self.email = input('Por favor, para receber o relatorio no se email, digite Seu Email:\n').lower()
        if not re.match(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', self.email):
            print('Please enter a valid email!!!')
            self.get_user_email()

    def scrape_data(self):
        #Aqui e exatamente onde o scrapper sera iniciado ( entrando no mercado livre) (Este método é responsável por acessar o site do Mercado Livre, realizar a busca pelo produto desejado e coletar as informações de preço e nome dos produtos. Ele faz isso iterando por várias páginas de resultados e buscando as informações relevantes no HTML de cada página.)
        url_base = 'https://lista.mercadolivre.com.br/'
        product_name = input('Qual Produto voce esta procurando? ') #Nessa parte vem o segundo prompt, do qual ira pedir para que voce escreva o produto selecionado


        page_number = 1
        products_per_page = 50

        while True:
                  url = f"{url_base}{product_name}_Desde_{page_number}"
                  response = requests.get(url)
                  site = BeautifulSoup(response.text, 'html.parser')

                  products = site.findAll('div', attrs={'class': 'ui-search-result__wrapper'})

                  if not products:
                     break

                  for product in products:
                      title = product.find('h2', attrs={'class': 'ui-search-item__title'})
                      if title:
                           self.product_name_list.append(title.text)

                      price_whole = product.find('span', attrs={'class': 'price-tag-fraction'})
                      price_decimal = product.find('span', attrs={'class': 'price-tag-cents'})

                      if price_whole and price_decimal:
                         self.product_price_list.append(f"R${price_whole.text},{price_decimal.text}")
                      elif price_whole:
                         self.product_price_list.append(f"R${price_whole.text}")

                  page_number += products_per_page


    def create_excel_file(self):
        #Nessa etapa sera criado o arquivo em Excel (Este método cria um arquivo Excel com os nomes e preços dos produtos encontrados. Ele usa a biblioteca openpyxl para isso.)
        index = 2
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Produtos'
        sheet['A1'] = 'Nome'
        sheet['B1'] = 'Valor'

        for name, price in zip(self.product_name_list, self.product_price_list):
            sheet.cell(column=1, row=index, value=name)
            sheet.cell(column=2, row=index, value=price)
            index += 1

        try:
            workbook.save("Buscador.xlsx")
            print('Arquivo em Excel, criado com sucesso!')
        except Exception as e:
            print(f'Falha ao criar o arquivo em Excel: {e}')

    def send_email(self):
        # Nessa etapa estamos enviando o email, com o arquivo em Excel em anexo. Esse email foi criado com único fim de utilizar esse script.
        sender_email = 'testeprogramador72@gmail.com'
        sender_password = 'vivtewcyalppxqoi'
        
        try:
            # Utilizando yagmail para enviar emails 
            yag = yagmail.SMTP(user=sender_email, password=sender_password)
            contents = [
                "Olá, sua planilha do Mercado Livre chegou.",
                "Buscador.xlsx"
            ]
            yag.send(self.email, 'Planilha Mercado Livre', contents)
            print('Email enviado com sucesso!')
        except Exception as e:
            print(f'Falha ao enviar o Email: {e}')

scrappy = Scrappy()
scrappy.start()

